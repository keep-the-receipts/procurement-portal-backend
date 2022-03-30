from unittest.mock import MagicMock, patch
import html5lib
import io
from django.contrib.messages import ERROR
from django.contrib.admin.sites import AdminSite
from django.test import Client, TestCase
from openpyxl import load_workbook
from .models import Dataset, DatasetVersion, PurchaseRecord, Repository
from .views import PURCHASE_RECORD_XLSX_FIELDS

DATASET_VERSION_VALUES = {
    "description": "test-description",
    "file": "dataset-uploads/Demo_dataset_1-2020-09-19T184114.321272.csv",
}
DATASET_VALUES = {
    "name": "test-dataset-name",
    "description": "test-dataset-description",
    "provenance": "test-dataset-provenance",
    "online_source_url": "test-dataset-online_source_url",
    "trusted_archive_url": "test-dataset-trusted_archive_url",
}

PURCHASE_RECORD_VALUES = {
    "supplier_name": "test-supplier_name",
    "order_amount_zar": 1,
    "invoice_amount_zar": 2,
    "payment_amount_zar": 3,
    "cost_per_unit_zar": 4,
    "buyer_name": "test-buyer_name",
    "central_supplier_database_number": "test-central_supplier_database_number",
    "company_registration_number": "test-company_registration_number",
    "director_names": "test-director_names",
    "director_names_and_surnames": "test-director_names_and_surnames",
    "director_surnames": "test-director_surnames",
    "implementation_location_district_municipality": "test-implementation_location_district_municipalityr",
    "implementation_location_facility": "test-implementation_location_facility",
    "implementation_location_local_municipality": "test-implementation_location_local_municipality",
    "implementation_location_other": "test-implementation_location_other",
    "implementation_location_province": "test-implementation_location_province",
    "items_description": "test-items_description",
    "items_quantity": "test-items_quantity",
    "items_unit": "test-items_unit",
    "procurement_method": "test-procurement_method",
    "state_employee": "test-state_employee",
    "award_date": "2020-01-01",
    "invoice_date": "2020-01-02",
    "invoice_receipt_date": "2020-01-03",
    "payment_date": "2020-01-04",
    "order_number": "test-order_number",
    "invoice_number": "test-invoice_number",
    "payment_number": "test-payment_number",
    "payment_period": "test-payment_period",
    "bbbee_status": "test-bbbee_status",
}

REPOSITORY_VALUES = {
    "name": "test-repository-name",
    "description": "test-repository-description",
}


class IndexTestCase(TestCase):
    def setUp(self):
        repository = Repository.objects.create(
            **REPOSITORY_VALUES
        )
        dataset = Dataset.objects.create(
            repository_id=repository.id,
            **DATASET_VALUES
        )
        dataset_version = DatasetVersion.objects.create(
            dataset=dataset,
            **DATASET_VERSION_VALUES
        )
        PurchaseRecord.objects.create(
            dataset_version=dataset_version,
            **PURCHASE_RECORD_VALUES
        )

    def test_index(self):
        c = Client()
        response = c.get("/records")
        self.assertContains(
            response,
            "index for records in procurement_portal",
        )
        assertValidHTML(response.content)

    def test_purchase_record_xlsx_streaming(self):
        c = Client()
        response = c.get("/records/purchase_records.xlsx?buyer_name=test-buyer_name")
        content = io.BytesIO(b"".join(response.streaming_content))
        workbook = load_workbook(content)
        worksheet = workbook.active
        for index, field in enumerate(PURCHASE_RECORD_XLSX_FIELDS):
            column = index + 1
            expected_column_heading = field.replace('__', '.')
            actual_column_heading = worksheet.cell(column=column, row=1).value
            self.assertEqual(
                expected_column_heading,
                actual_column_heading,
                'Column heading not as expected'
            )
            field_parts = expected_column_heading.split('.')
            if field_parts[0] != 'dataset_version':
                key = field_parts[0]
                expected_value = PURCHASE_RECORD_VALUES.get(key)
            elif field_parts[1] != 'dataset':
                key = field_parts[1]
                expected_value = DATASET_VERSION_VALUES.get(key)
            elif field_parts[2] != 'repository':
                key = field_parts[2]
                expected_value = DATASET_VALUES.get(key)
            else:
                key = field_parts[3]
                expected_value = REPOSITORY_VALUES.get(key)
            if isinstance(expected_value, int):
                expected_value = f"{expected_value}.00"
            if expected_value is None:
                print(f"Field {expected_column_heading} not checked")
            else:
                actual_value = worksheet.cell(column=column, row=2).value
                self.assertEqual(
                    expected_value,
                    actual_value,
                    f'Row value for column "{expected_column_heading}" not as expected'
                )
        next_column = len(PURCHASE_RECORD_XLSX_FIELDS) + 1
        next_column_heading = worksheet.cell(column=next_column, row=1).value
        self.assertEqual(
            None,
            next_column_heading,
            'Column heading not empty as expected'
        )
        next_column_value = worksheet.cell(column=next_column, row=2).value
        self.assertEqual(
            None,
            next_column_value,
            'Column value not empty as expected'
        )


def assertValidHTML(string):
    """
    Raises exception if the string is not valid HTML, e.g. has unmatched tags
    that need to be matched.
    """
    parser = html5lib.HTMLParser(strict=True)
    parser.parse(string)
