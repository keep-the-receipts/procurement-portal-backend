from django.apps import AppConfig
from openpyxl import Workbook


class RecordsConfig(AppConfig):
    name = "procurement_portal.records"

    def ready(self):
        import procurement_portal.records.signals  # noqa
        save_xlsx_template()


def save_xlsx_template():
    print("Saving template.xlsx")
    from .views import PURCHASE_RECORD_XLSX_FIELDS
    workbook = Workbook()
    worksheet = workbook.active
    for index, field in enumerate(PURCHASE_RECORD_XLSX_FIELDS):
        column = index + 1
        column_header = field.replace('__', '.')
        worksheet.cell(column=column, row=1).value = column_header
        worksheet.cell(column=column, row=2).value = 'text'
    workbook.save("template.xlsx")
    print("Saved template.xlsx")
